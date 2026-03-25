import io
from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import get_db
from app.models import Settlement, Campaign, Influencer
from app.models.user import User
from app.auth.dependencies import get_current_user

router = APIRouter(prefix="/settlements")
templates = Jinja2Templates(directory="app/templates")

SELLER_TYPES = ["사업자", "간이사업자", "프리랜서"]


def calc_settlement(sales_amount: float, commission_rate: float, seller_type: str) -> dict:
    """
    사업자    : 커미션만 지급 (부가세·원천징수 없음)
               final = commission

    간이사업자 : 커미션 - 부가세(10%)
               final = commission - vat

    프리랜서   : 커미션 합계 = 공급가액 + 부가세(10%)
               공급가액 = commission / 1.1
               원천징수 = 공급가액 × 3.3%   ← 부가세 제외 금액에 적용
               부가세는 세금계산서로 별도 처리 → 실지급 차감 안 함
               final = commission - 원천징수
    """
    commission = sales_amount * commission_rate
    if seller_type == "사업자":
        vat = 0
        withholding = 0
        tax_rate = 0.0
        final = round(commission)
    elif seller_type == "간이사업자":
        vat = round(commission * 0.1)
        withholding = 0
        tax_rate = 0.0
        final = round(commission) - vat
    else:  # 프리랜서
        supply_price = commission / 1.1            # 공급가액 (부가세 제외)
        vat = round(commission - supply_price)     # 부가세 (표시용)
        withholding = round(supply_price * 0.033)  # 원천징수 = 공급가액 × 3.3%
        tax_rate = 0.033
        final = round(commission) - withholding    # 부가세는 계산서 처리, 원천징수만 차감
    return {
        "commission_amount": round(commission),
        "vat_amount": vat,
        "tax_rate": tax_rate,
        "tax_amount": withholding,
        "final_payment": final,
    }


@router.get("")
def settlement_list(request: Request, db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user),
                    tab: str = "pending"):
    if tab not in ("pending", "confirmed", "paid"):
        tab = "pending"

    # Aggregate counts/amounts via SQL — no full table load
    total_paid = db.query(func.sum(Settlement.final_payment)).filter(
        Settlement.status == "paid"
    ).scalar() or 0
    pending_count = db.query(func.count(Settlement.id)).filter(
        Settlement.status == "pending"
    ).scalar() or 0
    confirmed_count = db.query(func.count(Settlement.id)).filter(
        Settlement.status == "confirmed"
    ).scalar() or 0
    paid_count = db.query(func.count(Settlement.id)).filter(
        Settlement.status == "paid"
    ).scalar() or 0

    # Sync seller_type from influencer.business_type if mismatched
    pending_settlements = db.query(Settlement).filter(
        Settlement.status.in_(["pending", "confirmed"]),
        Settlement.influencer_id.isnot(None),
    ).all()
    synced = False
    for st in pending_settlements:
        inf = st.influencer
        if inf and inf.business_type and inf.business_type != st.seller_type:
            calc = calc_settlement(st.sales_amount or 0, st.commission_rate or 0, inf.business_type)
            st.seller_type = inf.business_type
            st.vat_amount = calc["vat_amount"]
            st.tax_rate = calc["tax_rate"]
            st.tax_amount = calc["tax_amount"]
            st.commission_amount = calc["commission_amount"]
            st.final_payment = calc["final_payment"]
            synced = True
    if synced:
        db.commit()

    # Load only the active tab's rows
    settlements = (
        db.query(Settlement)
        .filter(Settlement.status == tab)
        .order_by(Settlement.created_at.desc())
        .limit(200)
        .all()
    )

    # Unique periods for export filter (distinct query, no full load)
    periods = sorted(
        {r[0] for r in db.query(Settlement.period_label)
         .filter(Settlement.period_label.isnot(None)).distinct().all()},
        reverse=True,
    )

    return templates.TemplateResponse("settlements/index.html", {
        "request": request, "active_page": "settlements", "current_user": current_user,
        "settlements": settlements, "total_paid": total_paid,
        "pending_count": pending_count, "confirmed_count": confirmed_count, "paid_count": paid_count,
        "tab": tab, "periods": periods,
        "seller_types": SELLER_TYPES,
    })


@router.get("/export")
def settlement_export(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    period: str = "",
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = db.query(Settlement)
    if period:
        query = query.filter(Settlement.period_label == period)
    rows = query.order_by(Settlement.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "정산내역"

    headers = ["기간", "인플루언서", "사업자유형", "캠페인", "매출액", "커미션율", "커미션액", "원천세율", "원천세액", "실지급액", "상태", "은행", "계좌번호", "예금주", "비고"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_labels = {"pending": "미확정", "confirmed": "확정됨", "paid": "지급완료"}

    for ri, s in enumerate(rows, 2):
        inf = s.influencer
        camp = s.campaign
        ws.cell(row=ri, column=1, value=s.period_label or "")
        ws.cell(row=ri, column=2, value=inf.name if inf else "")
        ws.cell(row=ri, column=3, value=s.seller_type or "")
        ws.cell(row=ri, column=4, value=camp.name if camp else "")
        ws.cell(row=ri, column=5, value=s.sales_amount or 0)
        ws.cell(row=ri, column=6, value=f"{(s.commission_rate or 0) * 100:.1f}%")
        ws.cell(row=ri, column=7, value=s.commission_amount or 0)
        ws.cell(row=ri, column=8, value=f"{(s.tax_rate or 0) * 100:.1f}%")
        ws.cell(row=ri, column=9, value=s.tax_amount or 0)
        ws.cell(row=ri, column=10, value=s.final_payment or 0)
        ws.cell(row=ri, column=11, value=status_labels.get(s.status, s.status))
        ws.cell(row=ri, column=12, value=inf.bank_name if inf else "")
        ws.cell(row=ri, column=13, value=inf.account_number if inf else "")
        ws.cell(row=ri, column=14, value=inf.account_holder if inf else "")
        ws.cell(row=ri, column=15, value=s.notes or "")

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    period_part = f"_{period}" if period else ""
    filename = quote(f"정산내역{period_part}.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/{settlement_id}")
def settlement_detail(settlement_id: str, request: Request,
                      db: Session = Depends(get_db),
                      current_user: User = Depends(get_current_user)):
    s = db.query(Settlement).filter(Settlement.id == settlement_id).first()
    if not s:
        return RedirectResponse("/settlements", status_code=302)
    return templates.TemplateResponse("settlements/detail.html", {
        "request": request, "active_page": "settlements",
        "current_user": current_user, "s": s,
    })


@router.post("/new")
def settlement_create(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    influencer_id: str = Form(""),
    campaign_id: str = Form(""),
    period_label: str = Form(""),
    seller_type: str = Form("사업자"),
    sales_amount: float = Form(0.0),
    commission_rate: float = Form(0.15),
    final_payment_manual: float = Form(0.0),
    notes: str = Form(""),
):
    calc = calc_settlement(sales_amount, commission_rate, seller_type)
    if final_payment_manual and final_payment_manual != calc["final_payment"]:
        calc["final_payment"] = round(final_payment_manual)
    s = Settlement(
        influencer_id=influencer_id or None,
        campaign_id=campaign_id or None,
        period_label=period_label or None,
        seller_type=seller_type,
        sales_amount=sales_amount,
        commission_rate=commission_rate,
        **calc,
        notes=notes or None,
    )
    db.add(s)
    db.commit()
    return RedirectResponse("/settlements?msg=정산+내역이+등록되었습니다", status_code=302)


@router.post("/{settlement_id}/recalc")
def settlement_recalc(
    settlement_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    seller_type: str = Form("사업자"),
    final_payment_manual: float = Form(0.0),
):
    """유형 변경 후 금액 재계산. final_payment_manual이 있으면 자동계산 대신 해당 값 사용."""
    s = db.query(Settlement).filter(Settlement.id == settlement_id).first()
    if not s:
        return RedirectResponse("/settlements", status_code=302)
    calc = calc_settlement(s.sales_amount or 0, s.commission_rate or 0, seller_type)
    s.seller_type = seller_type
    s.vat_amount = calc["vat_amount"]
    s.tax_rate = calc["tax_rate"]
    s.tax_amount = calc["tax_amount"]
    s.commission_amount = calc["commission_amount"]
    s.final_payment = round(final_payment_manual) if final_payment_manual else calc["final_payment"]
    db.commit()
    return RedirectResponse(f"/settlements/{settlement_id}?msg=저장+완료", status_code=302)


@router.post("/{settlement_id}/confirm")
def settlement_confirm(settlement_id: str, db: Session = Depends(get_db),
                       current_user: User = Depends(get_current_user)):
    s = db.query(Settlement).filter(Settlement.id == settlement_id).first()
    if s:
        s.status = "confirmed"
        db.commit()
    return RedirectResponse("/settlements?msg=확정되었습니다", status_code=302)


@router.post("/{settlement_id}/paid")
def settlement_paid(settlement_id: str, db: Session = Depends(get_db),
                    current_user: User = Depends(get_current_user)):
    s = db.query(Settlement).filter(Settlement.id == settlement_id).first()
    if s:
        s.status = "paid"
        db.commit()
    return RedirectResponse("/settlements?msg=지급+완료+처리되었습니다", status_code=302)


@router.post("/{settlement_id}/delete")
def settlement_delete(settlement_id: str, db: Session = Depends(get_db),
                      current_user: User = Depends(get_current_user)):
    s = db.query(Settlement).filter(Settlement.id == settlement_id).first()
    if s:
        db.delete(s)
        db.commit()
    return RedirectResponse("/settlements?msg=삭제되었습니다", status_code=302)
