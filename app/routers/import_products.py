"""
Excel / CSV product import router.
Routes:
  POST /products/import/upload   — parse file, return preview partial (htmx)
  POST /products/import/confirm  — save confirmed rows (with optional AI enrichment)
"""
import asyncio
import io
import json
import logging
import uuid
from typing import Optional

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, Form, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Product
from app.auth.dependencies import get_current_user
from app.auth.tenant import get_company_id
from app.models.user import User

router = APIRouter(prefix="/products/import")
templates = Jinja2Templates(directory="app/templates")

# ── Column name → Product field mapping ───────────────────────────────────────
COLUMN_MAP: dict[str, str] = {
    # 기본
    "제품명": "name", "상품명": "name",
    "브랜드": "brand",
    "카테고리": "category",
    "제품설명": "description", "설명": "description", "상품설명": "description",
    "상태": "status",
    "공개여부": "visibility_status", "카탈로그공개": "visibility_status", "카탈로그공개여부": "visibility_status",
    # 가격
    "소비자가": "consumer_price", "소비자가격": "consumer_price", "정가": "consumer_price",
    "최저가": "lowest_price",
    "공급가": "supplier_price", "원가": "supplier_price", "매입가": "supplier_price",
    "공구가": "groupbuy_price", "공동구매가": "groupbuy_price", "판매가": "groupbuy_price",
    "할인율": "discount_rate",
    "셀러커미션": "seller_commission_rate", "커미션율": "seller_commission_rate", "셀러커미션율": "seller_commission_rate",
    "벤더마진": "vendor_commission_rate", "벤더마진율": "vendor_commission_rate",
    "추천커미션율": "recommended_commission_rate",
    # 링크/이미지
    "상품링크": "product_link", "링크": "product_link", "제품링크": "product_link", "제품URL": "product_link",
    "이미지URL": "product_image", "이미지url": "product_image", "이미지링크": "product_image",
    "제품이미지": "product_image", "상품이미지": "product_image", "이미지": "product_image",
    # 마케팅
    "핵심셀링포인트": "unique_selling_point", "USP": "unique_selling_point", "usp": "unique_selling_point",
    "핵심혜택": "key_benefits", "혜택": "key_benefits",
    "콘텐츠앵글": "content_angle", "앵글": "content_angle", "콘텐츠각도": "content_angle",
    "포지셔닝전략": "positioning", "포지셔닝": "positioning",
    "공구가이드라인": "group_buy_guideline", "가이드라인": "group_buy_guideline",
    "카테고리태그": "categories", "소비자카테고리": "categories", "태그": "categories",
    # 배송
    "배송비": "shipping_cost",
    "배송유형": "shipping_type", "배송비유형": "shipping_type",
    "택배사": "carrier",
    "출고지": "ship_origin",
    "출고일": "dispatch_days", "발송기간": "dispatch_days",
    # 샘플
    "샘플여부": "sample_type", "샘플제공여부": "sample_type", "샘플": "sample_type",
    "샘플가격": "sample_price",
    # English
    "name": "name", "product_name": "name",
    "brand": "brand",
    "category": "category",
    "description": "description",
    "status": "status",
    "visibility_status": "visibility_status",
    "consumer_price": "consumer_price",
    "lowest_price": "lowest_price",
    "supplier_price": "supplier_price",
    "groupbuy_price": "groupbuy_price",
    "discount_rate": "discount_rate",
    "seller_commission_rate": "seller_commission_rate",
    "vendor_commission_rate": "vendor_commission_rate",
    "recommended_commission_rate": "recommended_commission_rate",
    "shipping_cost": "shipping_cost",
    "product_link": "product_link",
    "product_image": "product_image",
    "unique_selling_point": "unique_selling_point",
    "key_benefits": "key_benefits",
    "content_angle": "content_angle",
    "positioning": "positioning",
    "group_buy_guideline": "group_buy_guideline",
    "categories": "categories",
    "shipping_type": "shipping_type",
    "carrier": "carrier",
    "ship_origin": "ship_origin",
    "dispatch_days": "dispatch_days",
    "sample_type": "sample_type",
    "sample_price": "sample_price",
}

NUMERIC_FIELDS = {
    "consumer_price", "lowest_price", "supplier_price", "groupbuy_price",
    "shipping_cost", "sample_price",
}
PERCENT_FIELDS = {
    "discount_rate", "seller_commission_rate", "vendor_commission_rate",
    "recommended_commission_rate",
}
# 쉼표로 구분된 리스트 필드
LIST_FIELDS = {"key_benefits", "categories"}

ALL_PRODUCT_FIELDS = [
    ("name", "제품명 *"),
    ("brand", "브랜드 *"),
    ("category", "카테고리"),
    ("status", "상태"),
    ("visibility_status", "카탈로그 공개여부"),
    ("product_image", "이미지 URL"),
    ("product_link", "제품 URL"),
    ("description", "제품 설명"),
    ("consumer_price", "소비자가"),
    ("groupbuy_price", "공구가"),
    ("seller_commission_rate", "셀러 커미션율 (%)"),
    ("recommended_commission_rate", "추천 커미션율 (%)"),
    ("discount_rate", "할인율 (%)"),
    ("lowest_price", "최저가"),
    ("supplier_price", "공급가"),
    ("vendor_commission_rate", "벤더 마진율 (%)"),
    ("unique_selling_point", "핵심 셀링 포인트"),
    ("key_benefits", "핵심 혜택 (쉼표 구분)"),
    ("content_angle", "콘텐츠 앵글"),
    ("positioning", "포지셔닝 전략"),
    ("group_buy_guideline", "공구 가이드라인"),
    ("categories", "소비자 카테고리 태그 (쉼표 구분)"),
    ("shipping_type", "배송 유형"),
    ("shipping_cost", "배송비"),
    ("carrier", "택배사"),
    ("ship_origin", "출고지"),
    ("dispatch_days", "출고일"),
    ("sample_type", "샘플 제공 여부"),
    ("sample_price", "샘플 가격"),
    ("__skip__", "— 가져오지 않음 —"),
]


def _parse_file(content: bytes, filename: str) -> tuple[list[str], list[list]]:
    """Return (headers, rows) from xlsx or csv bytes."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = [[str(cell.value or "").strip() for cell in row] for row in ws.iter_rows()]
        wb.close()
    else:
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [[cell.strip() for cell in r] for r in reader]

    if not rows:
        return [], []
    headers = rows[0]
    data = [r for r in rows[1:] if any(c for c in r)]
    # 블랜드펀치 템플릿 힌트행 감지 후 제거 (2행에 "필수" 또는 "draft/active/archived" 포함)
    if data and any(cell in ("필수", "draft/active/archived", "active/hidden") for cell in data[0]):
        data = data[1:]
    return headers, data


def _auto_map(headers: list[str]) -> dict[int, str]:
    """Return {col_index: field_name}. Unknown cols map to '__skip__'."""
    mapping = {}
    for i, h in enumerate(headers):
        key = h.strip().lower().replace(" ", "").replace("_", "")
        matched = None
        for alias, field in COLUMN_MAP.items():
            if alias.lower().replace(" ", "").replace("_", "") == key:
                matched = field
                break
        mapping[i] = matched or "__skip__"
    return mapping


def _convert_value(field: str, raw: str):
    raw = raw.strip()
    if not raw or raw.lower() in ("none", "null", "-", "n/a"):
        return None
    if field in NUMERIC_FIELDS:
        try:
            return float(raw.replace(",", ""))
        except ValueError:
            return None
    if field in PERCENT_FIELDS:
        try:
            v = float(raw.replace(",", "").replace("%", ""))
            return v / 100.0 if v > 1 else v  # accept both 30 and 0.30
        except ValueError:
            return None
    if field in LIST_FIELDS:
        items = [s.strip() for s in raw.replace("，", ",").split(",") if s.strip()]
        return items if items else None
    return raw or None


def _ai_enrich(name: str, brand: str, category: str, description: str) -> dict | None:
    """Call Claude to extract structured fields from product description."""
    try:
        from pathlib import Path
        from app.ai.client import ClaudeClient

        prompt_template = Path("app/prompts/product_import_fill.md").read_text()
        prompt = (
            prompt_template
            .replace("{{name}}", name)
            .replace("{{brand}}", brand)
            .replace("{{category}}", category)
            .replace("{{description}}", description[:2000])
        )
        # Split on ## User Template
        parts = prompt.split("## User Template")
        system = parts[0].replace("## System\n", "").strip()
        user = parts[1].strip() if len(parts) > 1 else description

        client = ClaudeClient()
        result = client.complete_json(system=system, user=user)
        return result
    except Exception:
        return None


# ── Routes ────────────────────────────────────────────────────────────────────

@router.post("/upload", response_class=HTMLResponse)
async def import_upload(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    try:
        headers, rows = await asyncio.to_thread(_parse_file, content, file.filename or "file.xlsx")
    except Exception as e:
        return HTMLResponse(f'<p class="text-red-500 text-sm">파일 파싱 오류: {e}</p>')

    if not headers:
        return HTMLResponse('<p class="text-red-500 text-sm">파일이 비어 있습니다.</p>')

    mapping = _auto_map(headers)
    preview_rows = rows[:10]

    # Encode data for hidden field
    payload = json.dumps({
        "headers": headers,
        "rows": rows,
    })

    # Check if any column mapped to description (AI enrichment possible)
    has_description_col = any(v == "description" for v in mapping.values())

    return templates.TemplateResponse(
        "products/import_preview.html",
        {
            "request": request,
            "headers": headers,
            "preview_rows": preview_rows,
            "total_rows": len(rows),
            "mapping": mapping,
            "all_fields": ALL_PRODUCT_FIELDS,
            "payload": payload,
            "filename": file.filename,
            "has_description_col": has_description_col,
        },
    )


@router.post("/confirm")
async def import_confirm(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    payload_json: str = Form(...),
    mapping_json: str = Form(...),
    use_ai: str = Form(""),
):
    from urllib.parse import quote
    try:
        payload = json.loads(payload_json)
    except Exception as e:
        logger.exception("payload_json 파싱 실패")
        return RedirectResponse(f"/products/import?err={quote(f'payload JSON 오류: {e}')}", status_code=302)

    try:
        mapping = json.loads(mapping_json)
    except Exception as e:
        logger.exception("mapping_json 파싱 실패")
        return RedirectResponse(f"/products/import?err={quote(f'mapping JSON 오류: {e}')}", status_code=302)

    cid = get_company_id(current_user)
    headers = payload.get("headers", [])
    rows = payload.get("rows", [])
    run_ai = use_ai == "1"

    saved = 0
    ai_enriched = 0
    skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows):
        row_data: dict = {}
        for i, header in enumerate(headers):
            field = mapping.get(str(i), "__skip__")
            if field == "__skip__" or not field:
                continue
            val = row[i] if i < len(row) else ""
            converted = _convert_value(field, val)
            if converted is not None:
                row_data[field] = converted

        if not row_data.get("name"):
            skipped += 1
            continue  # skip rows without a product name

        # Optional AI enrichment from description
        ai_data = None
        if run_ai and row_data.get("description"):
            ai_data = await asyncio.to_thread(
                _ai_enrich,
                row_data.get("name", ""),
                row_data.get("brand", ""),
                row_data.get("category", ""),
                row_data["description"],
            )
            if ai_data:
                ai_enriched += 1

        # VARCHAR(20) 필드 검증 — 허용값 외 입력은 기본값으로 fallback
        _status = row_data.get("status", "draft")
        if _status not in ("draft", "active", "archived"):
            _status = "draft"
        _vis = row_data.get("visibility_status", "active")
        if _vis not in ("active", "hidden"):
            _vis = "active"

        def _trunc20(v):
            return v[:20] if isinstance(v, str) and len(v) > 20 else v

        try:
            product = Product(
                company_id=cid,
                name=row_data.get("name", ""),
                brand=row_data.get("brand", ""),
                category=row_data.get("category", "기타"),
                description=row_data.get("description"),
                status=_status,
                visibility_status=_vis,
                # 가격
                consumer_price=row_data.get("consumer_price", 0.0),
                lowest_price=row_data.get("lowest_price", 0.0),
                supplier_price=row_data.get("supplier_price", 0.0),
                groupbuy_price=row_data.get("groupbuy_price", 0.0),
                discount_rate=row_data.get("discount_rate", 0.0),
                seller_commission_rate=row_data.get("seller_commission_rate", 0.0),
                vendor_commission_rate=row_data.get("vendor_commission_rate", 0.0),
                recommended_commission_rate=row_data.get("recommended_commission_rate", 0.15),
                # 링크/이미지
                product_link=row_data.get("product_link"),
                product_image=row_data.get("product_image"),
                # 마케팅
                unique_selling_point=row_data.get("unique_selling_point") or (ai_data.get("unique_selling_point") if ai_data else None),
                key_benefits=row_data.get("key_benefits") or (ai_data.get("key_benefits") if ai_data else None),
                content_angle=row_data.get("content_angle") or (ai_data.get("content_angle") if ai_data else None),
                positioning=row_data.get("positioning"),
                group_buy_guideline=row_data.get("group_buy_guideline"),
                categories=row_data.get("categories"),
                target_audience=ai_data.get("target_audience") if ai_data else None,
                ai_analysis_raw=json.dumps(ai_data, ensure_ascii=False) if ai_data else None,
                # 배송
                shipping_cost=row_data.get("shipping_cost"),
                shipping_type=_trunc20(row_data.get("shipping_type")),
                carrier=_trunc20(row_data.get("carrier")),
                ship_origin=_trunc20(row_data.get("ship_origin")),
                dispatch_days=_trunc20(row_data.get("dispatch_days")),
                # 샘플
                sample_type=_trunc20(row_data.get("sample_type")),
                sample_price=row_data.get("sample_price"),
            )
            db.add(product)
            saved += 1
        except Exception as e:
            logger.exception(f"Row {row_idx + 2} 저장 오류: {row_data.get('name', '?')}")
            errors.append(f"행 {row_idx + 2} ({row_data.get('name', '?')}): {e}")
            continue

        # Commit in batches of 50
        if saved % 50 == 0:
            try:
                db.flush()
            except Exception as e:
                logger.exception("DB flush 오류")
                db.rollback()
                errors.append(f"DB flush 오류: {e}")

    try:
        db.commit()
    except Exception as e:
        logger.exception("DB commit 오류")
        db.rollback()
        from urllib.parse import quote
        return RedirectResponse(f"/products/import?err={quote(str(e))}", status_code=302)

    from urllib.parse import quote
    msg_parts = [f"{saved}개 제품 등록"]
    if run_ai and ai_enriched:
        msg_parts.append(f"AI 분석 {ai_enriched}개")
    if skipped:
        msg_parts.append(f"{skipped}개 건너뜀")
    if errors:
        msg_parts.append(f"오류 {len(errors)}행: {errors[0]}")
    return RedirectResponse(f"/products?msg={quote(' · '.join(msg_parts))}", status_code=302)


@router.get("/template")
def download_template(current_user: User = Depends(get_current_user)):
    """샘플 엑셀 템플릿 다운로드."""
    import io as _io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "제품목록"

    headers = [
        ("제품명", "필수"), ("브랜드", "필수"), ("카테고리", "기타"),
        ("상태", "draft/active/archived"), ("카탈로그공개여부", "active/hidden"),
        ("이미지URL", "https://..."), ("제품URL", "https://..."),
        ("제품설명", ""), ("소비자가", "숫자"), ("공구가", "숫자"),
        ("셀러커미션율", "숫자(%)"), ("할인율", "숫자(%)"),
        ("핵심셀링포인트", ""), ("핵심혜택", "쉼표로 구분"),
        ("콘텐츠앵글", ""), ("포지셔닝전략", ""),
        ("배송유형", "무료배송/유료배송"), ("배송비", "숫자"),
        ("택배사", "CJ대한통운 등"), ("출고지", "국내/해외"),
        ("샘플여부", "무상/유상/없음"), ("소비자카테고리", "쉼표로 구분"),
    ]

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1E40AF")
    hint_fill = PatternFill("solid", fgColor="EFF6FF")
    req_fill = PatternFill("solid", fgColor="DC2626")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    hint_font = Font(color="3B82F6", size=9, italic=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (col_name, hint) in enumerate(headers, 1):
        # 헤더 행
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = req_fill if hint in ("필수",) else header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        # 힌트 행
        hint_cell = ws.cell(row=2, column=col_idx, value=hint)
        hint_cell.fill = hint_fill
        hint_cell.font = hint_font
        hint_cell.alignment = center
        hint_cell.border = border

    # 샘플 데이터 3행
    samples = [
        ["데켓 쿡플레이트 세트", "데켓", "주방용품", "active", "active",
         "https://example.com/img.jpg", "https://smartstore.naver.com/...",
         "인덕션 호환 올스텐 쿡플레이트", 89000, 69000,
         15, 22, "주방을 바꾸는 단 하나의 제품",
         "인덕션 호환, 올스텐 소재, 설거지 편리",
         "주방을 업그레이드하고 싶은 신혼부부", "프리미엄 주방용품",
         "유료배송", 3000, "CJ대한통운", "국내", "무상", "주방,살림,요리"],
        ["디어커스 수분크림", "디어커스", "스킨케어", "active", "active",
         "", "", "피부 장벽 강화 수분 크림", 38000, 29000,
         20, 24, "하루 종일 촉촉한 피부",
         "수분 24시간 지속, 저자극, 민감성 가능",
         "피부 고민을 해결하고 싶은 20-30대", "민감성 전문 스킨케어",
         "무료배송", 0, "한진택배", "국내", "유상", "뷰티,스킨케어"],
        ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
         "", "", "", "", "", ""],
    ]
    for r_idx, row in enumerate(samples, 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # 컬럼 너비
    widths = [18, 12, 12, 10, 12, 30, 30, 30, 10, 10,
              10, 10, 25, 30, 25, 25, 12, 10, 12, 10, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=blendpunch_product_template.xlsx"},
    )
