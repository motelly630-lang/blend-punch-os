"""
어드민 주문 관리
"""
import csv
import io
import logging
import os
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.order import Order
from app.models.sales_page import SalesPage
from app.models.product import Product
from app.models.seller import Seller
from app.models.business_info import BusinessInfo
from app.auth.dependencies import get_current_user
from app.auth.tenant import get_company_id
from app.models.user import User
from app.services.kakao_notify import send_kakao_shipping

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/orders")
templates = Jinja2Templates(directory="app/templates")


# ── 공개 API (SHOP → OS) ────────────────────────────────────────────────────

@router.post("/api/create")
async def order_create_api(request: Request, db: Session = Depends(get_db)):
    """결제 완료 후 blend-pick에서 호출하는 주문 생성 API"""
    import random, string
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "invalid body"}, status_code=400)

    secret_key = os.environ.get("TOSS_SECRET_KEY", "")
    is_test = secret_key.startswith("test_")

    date_str = datetime.now(timezone.utc).strftime("%Y%m%d")
    suffix = "".join(random.choices(string.digits, k=6))
    order_number = f"BP-{date_str}-{suffix}"

    order = Order(
        order_number=order_number,
        product_id=body.get("product_id"),
        sales_page_id=body.get("sales_page_id"),
        customer_name=body.get("customer_name"),
        customer_phone=body.get("customer_phone"),
        customer_email=body.get("customer_email"),
        shipping_name=body.get("shipping_name"),
        shipping_phone=body.get("shipping_phone"),
        shipping_address=body.get("shipping_address"),
        shipping_address2=body.get("shipping_address2"),
        shipping_zipcode=body.get("shipping_zipcode"),
        shipping_memo=body.get("shipping_memo"),
        option_name=body.get("option_name"),
        quantity=body.get("quantity", 1),
        unit_price=body.get("unit_price", 0),
        total_price=body.get("total_price", 0),
        payment_key=body.get("payment_key"),
        payment_method=body.get("payment_method"),
        payment_status="paid",
        paid_at=datetime.now(timezone.utc),
        order_status="confirmed",
        is_test=is_test,
    )
    db.add(order)
    db.commit()
    return {"ok": True, "order_id": order.id, "order_number": order.order_number}


# ── 어드민 내부 헬퍼 ─────────────────────────────────────────────────────────

def _get_orders(db: Session, status: str = "", seller_code: str = "",
                page_id: str = "", search: str = "", company_id: int = 1):
    q = db.query(Order).filter(Order.company_id == company_id)
    if status == "new":
        q = q.filter(Order.payment_status == "paid", Order.order_status == "confirmed")
    elif status == "pending":
        q = q.filter(Order.payment_status == "pending")
    elif status == "shipping":
        q = q.filter(Order.order_status == "shipping")
    elif status == "delivered":
        q = q.filter(Order.order_status == "delivered")
    elif status == "cancelled":
        q = q.filter(Order.order_status == "cancelled")
    elif status == "paid":
        q = q.filter(Order.payment_status == "paid")
    if seller_code:
        q = q.filter(Order.seller_code == seller_code)
    if page_id:
        q = q.filter(Order.sales_page_id == page_id)
    if search:
        q = q.filter(
            Order.customer_name.contains(search) |
            Order.customer_phone.contains(search) |
            Order.order_number.contains(search)
        )
    return q.order_by(Order.created_at.desc()).all()


@router.get("")
def orders_list(
    request: Request,
    status: str = "",
    seller_code: str = "",
    page_id: str = "",
    search: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    cid = get_company_id(user)
    orders = _get_orders(db, status, seller_code, page_id, search, company_id=cid)
    pages = db.query(SalesPage).filter(SalesPage.company_id == cid).order_by(SalesPage.created_at.desc()).all()
    sellers = db.query(Seller).filter(Seller.company_id == cid, Seller.is_active == True).all()

    # 탭 카운트
    counts = {
        "all": db.query(Order).filter(Order.company_id == cid).count(),
        "new": db.query(Order).filter(Order.company_id == cid, Order.payment_status == "paid", Order.order_status == "confirmed").count(),
        "pending": db.query(Order).filter(Order.company_id == cid, Order.payment_status == "pending").count(),
        "shipping": db.query(Order).filter(Order.company_id == cid, Order.order_status == "shipping").count(),
        "delivered": db.query(Order).filter(Order.company_id == cid, Order.order_status == "delivered").count(),
        "cancelled": db.query(Order).filter(Order.company_id == cid, Order.order_status == "cancelled").count(),
    }
    biz = db.query(BusinessInfo).filter(BusinessInfo.id == 1).first()
    orders_banner = biz.orders_banner_image if biz else None
    return templates.TemplateResponse("orders/index.html", {
        "request": request, "orders": orders, "pages": pages,
        "sellers": sellers, "counts": counts,
        "cur_status": status, "cur_seller": seller_code,
        "cur_page": page_id, "search": search,
        "user": user, "active_page": "orders",
        "orders_banner_image": orders_banner,
    })


@router.get("/export")
def orders_export(
    status: str = "",
    seller_code: str = "",
    page_id: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return RedirectResponse("/orders?err=openpyxl+미설치", status_code=302)

    cid = get_company_id(user)
    orders = _get_orders(db, status, seller_code, page_id, company_id=cid)
    pages_map = {p.id: p for p in db.query(SalesPage).filter(SalesPage.company_id == cid).all()}
    products_map = {p.id: p for p in db.query(Product).filter(Product.company_id == cid).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주문목록"

    headers = ["주문번호", "주문일시", "상품명", "옵션", "수량", "결제금액",
               "구매자명", "연락처", "결제상태", "주문상태",
               "셀러코드", "배송지", "우편번호", "택배사", "송장번호"]
    header_fill = PatternFill(fill_type="solid", fgColor="1E40AF")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_kr = {
        "pending": "결제대기", "paid": "결제완료",
        "cancelled": "취소", "refunded": "환불",
    }
    order_kr = {
        "pending": "주문대기", "confirmed": "주문확인",
        "shipping": "배송중", "delivered": "배송완료",
        "cancelled": "취소",
    }

    for row, o in enumerate(orders, 2):
        page = pages_map.get(o.sales_page_id)
        product = products_map.get(o.product_id)
        ws.append([
            o.order_number,
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
            page.title if page and page.title else (product.name if product else ""),
            o.option_name or "",
            o.quantity,
            int(o.total_price),
            o.customer_name,
            o.customer_phone,
            status_kr.get(o.payment_status, o.payment_status),
            order_kr.get(o.order_status, o.order_status),
            o.seller_code or "",
            o.shipping_address + (" " + o.shipping_address2 if o.shipping_address2 else ""),
            o.shipping_zipcode,
            o.carrier_name or "",
            o.tracking_number or "",
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"orders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/bulk-ship")
async def orders_bulk_ship(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    CSV 일괄 송장 업로드.

    CSV 형식 (헤더 필수):
      order_number,courier,tracking_number
      BP-20260101-ABCDEF,CJ대한통운,123456789

    - order_number 또는 order_id 중 하나 필요
    - courier, tracking_number 필수
    - 인코딩: UTF-8 (BOM 포함 가능) 또는 CP949(EUC-KR)
    """
    # ── 파일 읽기 & 인코딩 처리 ─────────────────────────────────────────
    content = await file.read()
    if not content:
        return JSONResponse({"ok": False, "error": "파일이 비어있습니다."}, status_code=400)

    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return JSONResponse({"ok": False, "error": "파일 인코딩을 읽을 수 없습니다. UTF-8 또는 EUC-KR로 저장해주세요."}, status_code=400)

    # ── CSV 파싱 & 헤더 검증 ─────────────────────────────────────────────
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return JSONResponse({"ok": False, "error": "CSV 헤더가 없습니다."}, status_code=400)

    headers = {f.strip().lower() for f in reader.fieldnames}
    id_fields = {"order_id", "order_number"}
    required = {"courier", "tracking_number"}

    if not (headers & id_fields):
        return JSONResponse(
            {"ok": False, "error": "order_number 또는 order_id 컬럼이 필요합니다."},
            status_code=400,
        )
    missing = required - headers
    if missing:
        return JSONResponse(
            {"ok": False, "error": f"필수 컬럼 누락: {', '.join(sorted(missing))}"},
            status_code=400,
        )

    # ── 행별 처리 ────────────────────────────────────────────────────────
    processed, skipped = 0, 0
    errors: list[str] = []

    rows = list(reader)
    if not rows:
        return JSONResponse({"ok": False, "error": "데이터 행이 없습니다."}, status_code=400)

    for line_no, raw_row in enumerate(rows, start=2):
        row = {k.strip().lower(): (v or "").strip() for k, v in raw_row.items()}

        order_number = row.get("order_number", "")
        order_id = row.get("order_id", "")
        courier = row.get("courier", "")
        tracking = row.get("tracking_number", "")
        identifier = order_number or order_id

        # 빈 행 skip
        if not identifier:
            skipped += 1
            continue

        # 필수값 누락
        if not courier or not tracking:
            errors.append(f"행 {line_no} ({identifier}): courier 또는 tracking_number 값이 없습니다.")
            skipped += 1
            continue

        # 주문 조회 (order_number 우선)
        if order_number:
            order = db.query(Order).filter(Order.order_number == order_number).first()
        else:
            order = db.query(Order).filter(Order.id == order_id).first()

        if not order:
            errors.append(f"행 {line_no}: 주문을 찾을 수 없습니다 ({identifier})")
            skipped += 1
            continue

        # 배송 정보 업데이트
        order.carrier_name = courier
        order.tracking_number = tracking
        order.order_status = "shipping"
        order.shipped_at = datetime.utcnow()
        processed += 1

        # 카카오 알림톡 (실패해도 처리 계속)
        try:
            await send_kakao_shipping(order)
        except Exception as e:
            logger.warning("카카오 알림 실패 (주문 처리는 완료): %s", e)

    db.commit()

    return JSONResponse({
        "ok": True,
        "processed": processed,
        "skipped": skipped,
        "errors": errors,
    })


@router.get("/{order_id}")
def order_detail(order_id: str, request: Request,
                 db: Session = Depends(get_db),
                 user: User = Depends(get_current_user)):
    cid = get_company_id(user)
    order = db.query(Order).filter(Order.company_id == cid, Order.id == order_id).first()
    if not order:
        return RedirectResponse("/orders?err=주문을+찾을+수+없습니다", status_code=302)
    page = db.query(SalesPage).filter(SalesPage.id == order.sales_page_id).first()
    product = db.query(Product).filter(Product.id == order.product_id).first()
    seller = db.query(Seller).filter(Seller.id == order.seller_id).first() if order.seller_id else None
    return templates.TemplateResponse("orders/detail.html", {
        "request": request, "order": order, "page": page,
        "product": product, "seller": seller,
        "user": user, "active_page": "orders",
    })


@router.post("/{order_id}/ship")
async def order_ship(
    order_id: str,
    carrier_name: str = Form(""),
    tracking_number: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    cid = get_company_id(user)
    order = db.query(Order).filter(Order.company_id == cid, Order.id == order_id).first()
    if order:
        order.order_status = "shipping"
        order.carrier_name = carrier_name or None
        order.tracking_number = tracking_number or None
        order.shipped_at = datetime.utcnow()
        db.commit()
        try:
            await send_kakao_shipping(order)
        except Exception as e:
            logger.warning("카카오 알림 실패: %s", e)
    return RedirectResponse(f"/orders/{order_id}?msg=배송처리됨", status_code=302)


@router.post("/{order_id}/deliver")
def order_deliver(order_id: str, db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    cid = get_company_id(user)
    order = db.query(Order).filter(Order.company_id == cid, Order.id == order_id).first()
    if order:
        order.order_status = "delivered"
        db.commit()
    return RedirectResponse(f"/orders/{order_id}?msg=배송완료처리됨", status_code=302)


@router.post("/{order_id}/cancel")
def order_cancel(
    order_id: str,
    notes: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    cid = get_company_id(user)
    order = db.query(Order).filter(Order.company_id == cid, Order.id == order_id).first()
    if order:
        order.order_status = "cancelled"
        order.payment_status = "cancelled"
        if notes:
            order.notes = notes
        db.commit()
    return RedirectResponse(f"/orders/{order_id}?msg=취소처리됨", status_code=302)


@router.post("/{order_id}/notes")
def order_notes(
    order_id: str,
    notes: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    cid = get_company_id(user)
    order = db.query(Order).filter(Order.company_id == cid, Order.id == order_id).first()
    if order:
        order.notes = notes or None
        db.commit()
    return RedirectResponse(f"/orders/{order_id}?msg=메모저장됨", status_code=302)
