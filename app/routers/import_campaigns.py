"""
Excel / CSV campaign import router.
  GET  /campaigns/import         — upload page
  POST /campaigns/import/upload  — parse file, return preview partial
  POST /campaigns/import/confirm — save confirmed rows
"""
import io
import json
import logging
from datetime import date
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.auth.dependencies import get_current_user
from app.database import get_db
from app.models import Campaign, Product, Influencer
from app.models.user import User

router = APIRouter(prefix="/campaigns/import")
templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)

COLUMN_MAP: dict[str, str] = {
    "캠페인명": "name", "캠페인이름": "name", "이름": "name",
    "제품명": "product_name", "제품": "product_name",
    "인플루언서": "influencer_name", "인플루언서명": "influencer_name",
    "시작일": "start_date", "시작날짜": "start_date",
    "종료일": "end_date", "종료날짜": "end_date",
    "예상판매량": "expected_sales", "예상판매": "expected_sales",
    "실판매량": "actual_sales", "실제판매량": "actual_sales",
    "실매출": "actual_revenue", "실제매출": "actual_revenue", "매출": "actual_revenue",
    "단가": "unit_price",
    "셀러커미션": "seller_commission_rate", "셀러커미션율": "seller_commission_rate",
    "벤더마진": "vendor_commission_rate", "벤더마진율": "vendor_commission_rate",
    "메모": "notes",
    # English
    "name": "name", "campaign": "name",
    "product": "product_name",
    "influencer": "influencer_name",
    "start_date": "start_date", "start": "start_date",
    "end_date": "end_date", "end": "end_date",
    "expected_sales": "expected_sales",
    "actual_sales": "actual_sales",
    "actual_revenue": "actual_revenue", "revenue": "actual_revenue",
    "unit_price": "unit_price",
    "seller_commission_rate": "seller_commission_rate",
    "vendor_commission_rate": "vendor_commission_rate",
    "notes": "notes",
}

NUMERIC_FIELDS = {"expected_sales", "actual_sales", "actual_revenue", "unit_price"}
PERCENT_FIELDS = {"seller_commission_rate", "vendor_commission_rate"}

ALL_FIELDS = [
    ("name", "캠페인명"),
    ("product_name", "제품명 (이름 매칭)"),
    ("influencer_name", "인플루언서명 (이름 매칭)"),
    ("start_date", "시작일 (YYYY-MM-DD)"),
    ("end_date", "종료일 (YYYY-MM-DD)"),
    ("expected_sales", "예상 판매량"),
    ("actual_sales", "실 판매량"),
    ("actual_revenue", "실 매출"),
    ("unit_price", "단가"),
    ("seller_commission_rate", "셀러 커미션율 (%)"),
    ("vendor_commission_rate", "벤더 마진율 (%)"),
    ("notes", "메모"),
    ("__skip__", "— 가져오지 않음 —"),
]


def _parse_file(content: bytes, filename: str) -> tuple[list[str], list[list]]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = [[str(cell.value or "").strip() for cell in row] for row in ws.iter_rows()]
        wb.close()
    else:
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        rows = [[c.strip() for c in r] for r in csv.reader(io.StringIO(text))]
    if not rows:
        return [], []
    return rows[0], [r for r in rows[1:] if any(c for c in r)]


def _auto_map(headers: list[str]) -> dict[int, str]:
    mapping = {}
    for i, h in enumerate(headers):
        key = h.strip().lower().replace(" ", "").replace("_", "")
        matched = None
        for alias, field in COLUMN_MAP.items():
            if alias.lower().replace(" ", "").replace("_", "") == key:
                matched = field
                break
        mapping[i] = matched or "__skip__"
    return mapping


def _convert(field: str, raw: str):
    raw = raw.strip()
    if not raw or raw.lower() in ("none", "null", "-", "n/a"):
        return None
    if field in NUMERIC_FIELDS:
        try:
            return float(raw.replace(",", ""))
        except ValueError:
            return None
    if field in PERCENT_FIELDS:
        try:
            v = float(raw.replace(",", "").replace("%", ""))
            return v / 100.0 if v > 1 else v
        except ValueError:
            return None
    if field in ("start_date", "end_date"):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
            try:
                return date.fromisoformat(raw) if fmt == "%Y-%m-%d" else __import__("datetime").datetime.strptime(raw, fmt).date()
            except Exception:
                continue
        return None
    return raw or None


# ── Routes ─────────────────────────────────────────────────────────────────────

@router.get("")
def import_page(request: Request, current_user: User = Depends(get_current_user)):
    return templates.TemplateResponse("campaigns/import.html", {
        "request": request, "active_page": "campaigns", "current_user": current_user,
    })


@router.post("/upload", response_class=HTMLResponse)
async def import_upload(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    try:
        import asyncio
        headers, rows = await asyncio.to_thread(_parse_file, content, file.filename or "file.xlsx")
    except Exception as e:
        return HTMLResponse(f'<p class="text-red-500 text-sm">파일 파싱 오류: {e}</p>')
    if not headers:
        return HTMLResponse('<p class="text-red-500 text-sm">파일이 비어 있습니다.</p>')

    mapping = _auto_map(headers)
    payload = json.dumps({"headers": headers, "rows": rows})
    return templates.TemplateResponse("campaigns/import_preview.html", {
        "request": request,
        "headers": headers,
        "preview_rows": rows[:10],
        "total_rows": len(rows),
        "mapping": mapping,
        "all_fields": ALL_FIELDS,
        "payload": payload,
        "filename": file.filename,
    })


@router.post("/confirm")
async def import_confirm(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    payload_json: str = Form(...),
    mapping_json: str = Form(...),
):
    try:
        payload = json.loads(payload_json)
    except Exception as e:
        return RedirectResponse(f"/campaigns/import?err={quote(f'payload 오류: {e}')}", status_code=302)
    try:
        mapping = json.loads(mapping_json)
    except Exception as e:
        return RedirectResponse(f"/campaigns/import?err={quote(f'mapping 오류: {e}')}", status_code=302)

    headers = payload.get("headers", [])
    rows = payload.get("rows", [])

    # Pre-load product/influencer name→id maps for matching
    product_map = {p.name.strip().lower(): p.id for p in db.query(Product).all()}
    inf_map = {i.name.strip().lower(): i.id for i in db.query(Influencer).all()}

    saved = skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows):
        row_data: dict = {}
        for i, _ in enumerate(headers):
            field = mapping.get(str(i), "__skip__")
            if field == "__skip__" or not field:
                continue
            val = row[i] if i < len(row) else ""
            converted = _convert(field, val)
            if converted is not None:
                row_data[field] = converted

        if not row_data.get("name"):
            skipped += 1
            continue

        # Resolve product/influencer by name
        product_id = None
        if row_data.get("product_name"):
            product_id = product_map.get(row_data["product_name"].strip().lower())

        influencer_id = None
        if row_data.get("influencer_name"):
            influencer_id = inf_map.get(row_data["influencer_name"].strip().lower())

        seller_rate = row_data.get("seller_commission_rate", 0.0) or 0.0
        vendor_rate = row_data.get("vendor_commission_rate", 0.0) or 0.0
        revenue = row_data.get("actual_revenue", 0.0) or 0.0

        try:
            c = Campaign(
                name=row_data["name"],
                product_id=product_id,
                influencer_id=influencer_id,
                start_date=row_data.get("start_date"),
                end_date=row_data.get("end_date"),
                expected_sales=int(row_data.get("expected_sales") or 0),
                actual_sales=int(row_data.get("actual_sales") or 0),
                actual_revenue=revenue,
                unit_price=row_data.get("unit_price", 0.0) or 0.0,
                seller_commission_rate=seller_rate,
                vendor_commission_rate=vendor_rate,
                seller_commission_amount=round(revenue * seller_rate),
                vendor_commission_amount=round(revenue * vendor_rate),
                notes=row_data.get("notes"),
                status="planning",
            )
            db.add(c)
            saved += 1
        except Exception as e:
            logger.exception(f"Row {row_idx + 2} 저장 오류")
            errors.append(f"행 {row_idx + 2}: {e}")

        if saved % 50 == 0:
            try:
                db.flush()
            except Exception as e:
                db.rollback()
                errors.append(f"DB flush 오류: {e}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return RedirectResponse(f"/campaigns/import?err={quote(str(e))}", status_code=302)

    parts = [f"{saved}개 캠페인 등록"]
    if skipped:
        parts.append(f"{skipped}개 건너뜀")
    if errors:
        parts.append(f"오류 {len(errors)}행")
    return RedirectResponse(f"/campaigns?msg={quote(' · '.join(parts))}", status_code=302)
