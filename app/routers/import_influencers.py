"""
Excel / CSV influencer import router.
Routes:
  GET  /influencers/import          — upload page
  POST /influencers/import/upload   — parse file, return preview partial (htmx)
  POST /influencers/import/confirm  — save confirmed rows
"""
import io
import json
import logging
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from app.auth.dependencies import get_current_user
from app.database import get_db
from app.models import Influencer
from app.models.user import User

router = APIRouter(prefix="/influencers/import")
templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)

# ── Column mapping ─────────────────────────────────────────────────────────────
COLUMN_MAP: dict[str, str] = {
    # Korean
    "이름": "name", "인플루언서명": "name", "활동명": "name",
    "플랫폼": "platform",
    "아이디": "handle", "핸들": "handle", "계정": "handle",
    "팔로워": "followers", "팔로워수": "followers",
    "카테고리": "categories",
    "참여율": "engagement_rate", "인게이지먼트": "engagement_rate",
    "프로필URL": "profile_url", "프로필링크": "profile_url",
    "연락이메일": "contact_email", "이메일": "contact_email",
    "연락전화": "contact_phone", "전화번호": "contact_phone", "전화": "contact_phone",
    "카카오": "contact_kakao", "카카오아이디": "contact_kakao",
    "소속사": "agency_name", "에이전시": "agency_name",
    "평균조회수": "avg_views_per_post",
    "과거매출": "past_gmv",
    "메모": "notes",
    # English
    "name": "name",
    "platform": "platform",
    "handle": "handle",
    "followers": "followers",
    "categories": "categories",
    "engagement_rate": "engagement_rate",
    "profile_url": "profile_url",
    "contact_email": "contact_email",
    "contact_phone": "contact_phone",
    "contact_kakao": "contact_kakao",
    "agency_name": "agency_name",
    "avg_views": "avg_views_per_post",
    "notes": "notes",
}

NUMERIC_INT_FIELDS = {"followers", "avg_views_per_post"}
NUMERIC_FLOAT_FIELDS = {"engagement_rate", "past_gmv"}

ALL_FIELDS = [
    ("name", "이름"),
    ("platform", "플랫폼"),
    ("handle", "아이디/핸들"),
    ("followers", "팔로워수"),
    ("categories", "카테고리 (쉼표 구분)"),
    ("engagement_rate", "참여율 (%)"),
    ("profile_url", "프로필 URL"),
    ("contact_email", "연락 이메일"),
    ("contact_phone", "연락 전화"),
    ("contact_kakao", "카카오 아이디"),
    ("agency_name", "소속사"),
    ("avg_views_per_post", "평균 조회수"),
    ("past_gmv", "과거 매출"),
    ("notes", "메모"),
    ("__skip__", "— 가져오지 않음 —"),
]

PLATFORM_ALIASES = {
    "인스타그램": "instagram", "인스타": "instagram", "ig": "instagram",
    "유튜브": "youtube", "yt": "youtube",
    "틱톡": "tiktok", "tt": "tiktok",
    "블로그": "blog", "네이버": "naver",
}
VALID_PLATFORMS = {"instagram", "youtube", "tiktok", "blog", "naver"}


def _parse_file(content: bytes, filename: str) -> tuple[list[str], list[list]]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = [[str(cell.value or "").strip() for cell in row] for row in ws.iter_rows()]
        wb.close()
    else:
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        rows = [[c.strip() for c in r] for r in csv.reader(io.StringIO(text))]

    if not rows:
        return [], []
    headers = rows[0]
    data = [r for r in rows[1:] if any(c for c in r)]
    return headers, data


def _auto_map(headers: list[str]) -> dict[int, str]:
    mapping = {}
    for i, h in enumerate(headers):
        key = h.strip().lower().replace(" ", "").replace("_", "")
        matched = None
        for alias, field in COLUMN_MAP.items():
            if alias.lower().replace(" ", "").replace("_", "") == key:
                matched = field
                break
        mapping[i] = matched or "__skip__"
    return mapping


def _convert_value(field: str, raw: str):
    raw = raw.strip()
    if not raw or raw.lower() in ("none", "null", "-", "n/a"):
        return None
    if field in NUMERIC_INT_FIELDS:
        try:
            return int(float(raw.replace(",", "").replace("만", "0000").replace("k", "000").replace("K", "000")))
        except ValueError:
            return None
    if field in NUMERIC_FLOAT_FIELDS:
        try:
            v = float(raw.replace(",", "").replace("%", ""))
            return v / 100.0 if v > 1 else v
        except ValueError:
            return None
    if field == "platform":
        p = raw.lower().strip()
        return PLATFORM_ALIASES.get(p, p if p in VALID_PLATFORMS else "instagram")
    if field == "categories":
        cats = [c.strip() for c in raw.replace("/", ",").split(",") if c.strip()]
        return cats if cats else None
    return raw or None


# ── Routes ─────────────────────────────────────────────────────────────────────

@router.get("")
def import_page(request: Request, current_user: User = Depends(get_current_user)):
    return templates.TemplateResponse("influencers/import.html", {
        "request": request, "active_page": "influencers", "current_user": current_user,
    })


@router.post("/upload", response_class=HTMLResponse)
async def import_upload(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    try:
        import asyncio
        headers, rows = await asyncio.to_thread(_parse_file, content, file.filename or "file.xlsx")
    except Exception as e:
        return HTMLResponse(f'<p class="text-red-500 text-sm">파일 파싱 오류: {e}</p>')

    if not headers:
        return HTMLResponse('<p class="text-red-500 text-sm">파일이 비어 있습니다.</p>')

    mapping = _auto_map(headers)
    payload = json.dumps({"headers": headers, "rows": rows})

    return templates.TemplateResponse("influencers/import_preview.html", {
        "request": request,
        "headers": headers,
        "preview_rows": rows[:10],
        "total_rows": len(rows),
        "mapping": mapping,
        "all_fields": ALL_FIELDS,
        "payload": payload,
        "filename": file.filename,
    })


@router.post("/confirm")
async def import_confirm(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    payload_json: str = Form(...),
    mapping_json: str = Form(...),
):
    try:
        payload = json.loads(payload_json)
    except Exception as e:
        logger.exception("payload_json 파싱 실패")
        return RedirectResponse(f"/influencers/import?err={quote(f'payload 오류: {e}')}", status_code=302)
    try:
        mapping = json.loads(mapping_json)
    except Exception as e:
        logger.exception("mapping_json 파싱 실패")
        return RedirectResponse(f"/influencers/import?err={quote(f'mapping 오류: {e}')}", status_code=302)

    headers = payload.get("headers", [])
    rows = payload.get("rows", [])
    saved = skipped = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows):
        row_data: dict = {}
        for i, _ in enumerate(headers):
            field = mapping.get(str(i), "__skip__")
            if field == "__skip__" or not field:
                continue
            val = row[i] if i < len(row) else ""
            converted = _convert_value(field, val)
            if converted is not None:
                row_data[field] = converted

        if not row_data.get("name") and not row_data.get("handle"):
            skipped += 1
            continue

        name = row_data.get("name") or row_data.get("handle", "")
        handle = row_data.get("handle") or ""

        try:
            inf = Influencer(
                name=name,
                handle=handle,
                platform=row_data.get("platform", "instagram"),
                followers=row_data.get("followers", 0),
                engagement_rate=row_data.get("engagement_rate", 0.0),
                categories=row_data.get("categories"),
                profile_url=row_data.get("profile_url"),
                contact_email=row_data.get("contact_email"),
                contact_phone=row_data.get("contact_phone"),
                contact_kakao=row_data.get("contact_kakao"),
                agency_name=row_data.get("agency_name"),
                avg_views_per_post=row_data.get("avg_views_per_post", 0),
                past_gmv=row_data.get("past_gmv", 0.0),
                notes=row_data.get("notes"),
                status="active",
            )
            db.add(inf)
            saved += 1
        except Exception as e:
            logger.exception(f"Row {row_idx + 2} 저장 오류")
            errors.append(f"행 {row_idx + 2}: {e}")

        if saved % 50 == 0:
            try:
                db.flush()
            except Exception as e:
                db.rollback()
                errors.append(f"DB flush 오류: {e}")

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.exception("DB commit 오류")
        return RedirectResponse(f"/influencers/import?err={quote(str(e))}", status_code=302)

    parts = [f"{saved}명 등록"]
    if skipped:
        parts.append(f"{skipped}개 건너뜀")
    if errors:
        parts.append(f"오류 {len(errors)}행")
    return RedirectResponse(f"/influencers?msg={quote(' · '.join(parts))}", status_code=302)
