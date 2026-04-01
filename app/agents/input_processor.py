"""
입력 데이터 전처리기

이미지(상세페이지), 엑셀(가격표), 텍스트를 받아
에이전트 파이프라인에 전달할 통합 컨텍스트를 생성한다.
"""
import base64
import io
import json
from pathlib import Path
from typing import Optional

import anthropic
from app.config import settings

_client = anthropic.Anthropic(api_key=settings.anthropic_api_key)


def process_image(image_path: str) -> dict:
    """
    상품 이미지/상세페이지에서 핵심 정보 추출.
    Claude Vision을 사용해 OCR + 정보 추출을 동시에 처리.
    """
    path = Path(image_path)
    if not path.exists():
        return {"error": f"파일 없음: {image_path}"}

    suffix = path.suffix.lower()
    media_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                 ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif"}
    media_type = media_map.get(suffix, "image/jpeg")

    with open(image_path, "rb") as f:
        image_data = base64.standard_b64encode(f.read()).decode("utf-8")

    resp = _client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=1024,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media_type, "data": image_data},
                },
                {
                    "type": "text",
                    "text": (
                        "이 상품 이미지에서 다음 정보를 추출해 JSON으로만 반환하세요:\n"
                        "{\n"
                        '  "product_name": "제품명",\n'
                        '  "brand_name": "브랜드명",\n'
                        '  "key_features": ["특징1", "특징2"],\n'
                        '  "ingredients_or_components": "성분/구성",\n'
                        '  "usage_instructions": "사용법",\n'
                        '  "extracted_text": "이미지에서 추출된 전체 텍스트"\n'
                        "}"
                    ),
                },
            ],
        }],
    )

    text = resp.content[0].text.strip()
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    try:
        return {"source": "image", "data": json.loads(text)}
    except Exception:
        return {"source": "image", "data": {"extracted_text": text}}


def process_excel(excel_path: str) -> dict:
    """
    엑셀 파일에서 가격/옵션 구조 추출.
    openpyxl로 파싱 후 구조화.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl 미설치. uv pip install openpyxl"}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheets_data = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cleaned = [str(c).strip() if c is not None else "" for c in row]
                if any(cleaned):
                    rows.append(cleaned)
            if rows:
                sheets_data[sheet_name] = rows[:50]  # 최대 50행

        # Claude로 구조 분석
        summary_text = json.dumps(sheets_data, ensure_ascii=False)[:3000]
        resp = _client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": (
                    f"다음 엑셀 데이터를 분석해 JSON으로만 반환하세요:\n{summary_text}\n\n"
                    "출력 형식:\n"
                    "{\n"
                    '  "options": [{"name": "옵션명", "price": 0, "unit": "단위"}],\n'
                    '  "price_structure": {"소비자가": 0, "공급가": 0, "최저가": 0},\n'
                    '  "key_info": "핵심 요약"\n'
                    "}"
                ),
            }],
        )
        text = resp.content[0].text.strip()
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        try:
            return {"source": "excel", "data": json.loads(text)}
        except Exception:
            return {"source": "excel", "data": {"raw": sheets_data}}
    except Exception as e:
        return {"source": "excel", "error": str(e)}


def process_text(text: str) -> dict:
    """텍스트 입력 그대로 전달."""
    return {"source": "text", "data": {"raw_text": text}}


def build_initial_context(
    text: Optional[str] = None,
    image_path: Optional[str] = None,
    excel_path: Optional[str] = None,
    existing_data: Optional[dict] = None,
) -> dict:
    """
    모든 입력을 통합해 파이프라인 초기 컨텍스트를 생성.

    Returns:
        {
            "inputs": {...},   # 원본 입력 데이터
            "combined": str,   # 통합 요약 텍스트
        }
    """
    inputs = {}

    if existing_data:
        inputs["db_data"] = existing_data

    if text:
        inputs["text"] = process_text(text)

    if image_path:
        inputs["image"] = process_image(image_path)

    if excel_path:
        inputs["excel"] = process_excel(excel_path)

    return {
        "inputs": inputs,
        "combined": json.dumps(inputs, ensure_ascii=False),
    }
